#!/usr/bin/env python3
"""
Excel Report Generator
Generates multi-sheet Excel workbooks with test results.
Usage:
    python automation/utils/excel_reporter.py --suite auth_authorization --output automation/reports
    python automation/utils/excel_reporter.py --merge --input automation/all-artifacts --output automation/reports/Excel
"""

import argparse
import json
import os
import sys
from datetime import datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent.parent))

try:
    import openpyxl
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, GradientFill
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, Reference
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("WARNING: openpyxl not available — skipping Excel reports")


# ─────────────────────────────────────────────────────────────────────────────
# Style helpers
# ─────────────────────────────────────────────────────────────────────────────
def _hdr_style() -> dict:
    return {
        'font': Font(bold=True, color='FFFFFF', size=11),
        'fill': PatternFill(fill_type='solid', fgColor='1E3A5F'),
        'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
        'border': Border(
            bottom=Side(border_style='medium', color='334155'),
        )
    }

def _cell_style(bg: str = 'FFFFFF') -> dict:
    return {
        'fill': PatternFill(fill_type='solid', fgColor=bg),
        'alignment': Alignment(vertical='center', wrap_text=True),
        'border': Border(bottom=Side(border_style='thin', color='E2E8F0')),
    }

def _apply(cell, style: dict):
    for attr, val in style.items():
        setattr(cell, attr, val)

STATUS_FILLS = {
    'Passed': 'D1FAE5',  # green tint
    'passed': 'D1FAE5',
    'PASS':   'D1FAE5',
    'Failed': 'FEE2E2',  # red tint
    'failed': 'FEE2E2',
    'FAIL':   'FEE2E2',
    'Error':  'FFEDD5',  # orange
    'error':  'FFEDD5',
    'Skipped':'FEF3C7',  # yellow
    'skipped':'FEF3C7',
    'SKIP':   'FEF3C7',
}

STATUS_FONTS = {
    'Passed': '166534', 'passed': '166534', 'PASS': '166534',
    'Failed': '991B1B', 'failed': '991B1B', 'FAIL': '991B1B',
    'Error':  '9A3412', 'error':  '9A3412',
    'Skipped':'92400E', 'skipped':'92400E', 'SKIP':'92400E',
}


def _load_results(input_path: str) -> list:
    results = []
    base = Path(input_path)
    for f in base.rglob('*.json'):
        try:
            with open(f, encoding='utf-8') as fh:
                data = json.load(fh)
            if isinstance(data, list):
                results.extend(data)
            elif isinstance(data, dict) and 'tests' in data:
                results.extend(data['tests'])
        except Exception:
            pass
    # Deduplicate
    seen, unique = set(), []
    for r in results:
        key = r.get('test_id', '') + r.get('name', '')
        if key not in seen:
            seen.add(key)
            unique.append(r)
    return unique


def _is_passed(r: dict) -> bool:
    return r.get('status', '').lower() in ('passed', 'pass')

def _is_failed(r: dict) -> bool:
    return r.get('status', '').lower() in ('failed', 'fail', 'error')

def _is_skipped(r: dict) -> bool:
    return r.get('status', '').lower() in ('skipped', 'skip')


def _write_test_rows(ws, results: list, row_offset: int = 2):
    COLS = ['Test ID', 'Module', 'Test Name', 'Status', 'Priority',
            'Execution Time', 'Error Message', 'Suite']
    widths = [14, 22, 60, 12, 10, 14, 50, 22]

    # Header row
    for col, (header, width) in enumerate(zip(COLS, widths), 1):
        cell = ws.cell(row=1, column=col, value=header)
        _apply(cell, _hdr_style())
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[1].height = 24

    for row_i, r in enumerate(results, row_offset):
        status = r.get('status', 'Unknown')
        bg = STATUS_FILLS.get(status, 'FFFFFF')
        fg = STATUS_FONTS.get(status, '1E293B')
        row_data = [
            r.get('test_id', r.get('name', '')[:20]),
            r.get('module', 'Unknown'),
            r.get('name', '')[:100],
            status,
            r.get('priority', 'Medium'),
            f"{r.get('execution_time_ms', 0):.0f}ms",
            (r.get('error_message', r.get('error', '')) or '')[:200],
            r.get('suite', ''),
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_i, column=col, value=val)
            cell.fill = PatternFill(fill_type='solid', fgColor=bg)
            cell.alignment = Alignment(vertical='center', wrap_text=True)
            cell.border = Border(bottom=Side(border_style='thin', color='E2E8F0'))
            if col == 4:  # Status column
                cell.font = Font(bold=True, color=fg)
        ws.row_dimensions[row_i].height = 18


def _add_metrics_sheet(wb, results: list):
    ws = wb.create_sheet('Execution Metrics')
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20

    total = len(results)
    passed = sum(1 for r in results if _is_passed(r))
    failed = sum(1 for r in results if _is_failed(r))
    skipped = sum(1 for r in results if _is_skipped(r))
    rate = round(passed / total * 100, 2) if total > 0 else 0

    ws.cell(1, 1, 'Metric').font = Font(bold=True, size=12, color='1E3A5F')
    ws.cell(1, 2, 'Value').font = Font(bold=True, size=12, color='1E3A5F')

    metrics = [
        ('Total Test Cases', total),
        ('Passed', passed),
        ('Failed', failed),
        ('Skipped', skipped),
        ('Pass Rate', f'{rate}%'),
        ('Execution Date', datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')),
        ('Framework', 'Selenium 4 + Python pytest'),
        ('Browser', 'Chrome (headless)'),
        ('Deployment', os.getenv('BASE_URL', 'GitHub Pages')),
    ]
    for row_i, (metric, value) in enumerate(metrics, 2):
        ws.cell(row_i, 1, metric).font = Font(bold=True)
        ws.cell(row_i, 2, str(value))
        if row_i % 2 == 0:
            for col in range(1, 3):
                ws.cell(row_i, col).fill = PatternFill(fill_type='solid', fgColor='F8FAFC')

    # Module breakdown
    ws.cell(14, 1, 'Module Breakdown').font = Font(bold=True, size=11, color='1E3A5F')
    modules: dict = {}
    for r in results:
        mod = r.get('module', 'Unknown')
        if mod not in modules:
            modules[mod] = {'total': 0, 'passed': 0, 'failed': 0}
        modules[mod]['total'] += 1
        if _is_passed(r): modules[mod]['passed'] += 1
        elif _is_failed(r): modules[mod]['failed'] += 1

    ws.cell(15, 1, 'Module').font = Font(bold=True)
    ws.cell(15, 2, 'Total').font = Font(bold=True)
    ws.cell(15, 3, 'Passed').font = Font(bold=True)
    ws.cell(15, 4, 'Failed').font = Font(bold=True)
    ws.cell(15, 5, 'Rate').font = Font(bold=True)
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12

    for row_i, (mod, ms) in enumerate(sorted(modules.items()), 16):
        rate_m = round(ms['passed'] / ms['total'] * 100, 1) if ms['total'] else 0
        ws.cell(row_i, 1, mod)
        ws.cell(row_i, 2, ms['total'])
        ws.cell(row_i, 3, ms['passed']).font = Font(color='166534')
        ws.cell(row_i, 4, ms['failed']).font = Font(color='991B1B')
        ws.cell(row_i, 5, f'{rate_m}%')


def _add_defect_summary(wb, failed_results: list):
    ws = wb.create_sheet('Defect Summary')
    ws.sheet_view.showGridLines = False
    for col, header in enumerate(['Test ID', 'Module', 'Test Name', 'Error Message', 'Priority'], 1):
        cell = ws.cell(1, col, header)
        _apply(cell, _hdr_style())
    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 55
    ws.column_dimensions['D'].width = 60
    ws.column_dimensions['E'].width = 12

    for row_i, r in enumerate(failed_results, 2):
        ws.cell(row_i, 1, r.get('test_id', ''))
        ws.cell(row_i, 2, r.get('module', ''))
        ws.cell(row_i, 3, r.get('name', '')[:100])
        ws.cell(row_i, 4, (r.get('error_message', r.get('error', '')) or '')[:200])
        ws.cell(row_i, 5, r.get('priority', 'Medium'))
        for col in range(1, 6):
            ws.cell(row_i, col).fill = PatternFill(fill_type='solid', fgColor='FEE2E2')
            ws.row_dimensions[row_i].height = 18


def generate_excel_reports(results: list, output_dir: str, suite: str = ''):
    """Generate all Excel report files."""
    if not HAS_OPENPYXL:
        print("Skipping Excel reports — openpyxl not installed")
        return

    os.makedirs(output_dir, exist_ok=True)
    passed = [r for r in results if _is_passed(r)]
    failed = [r for r in results if _is_failed(r)]
    skipped = [r for r in results if _is_skipped(r)]

    # ── 1. Main automation report ───────────────────────────────────────────
    wb = openpyxl.Workbook()
    wb.active.title = 'All Tests'
    _write_test_rows(wb.active, results)
    wb.active.freeze_panes = 'A2'

    passed_ws = wb.create_sheet('Passed Tests')
    _write_test_rows(passed_ws, passed)
    passed_ws.freeze_panes = 'A2'

    failed_ws = wb.create_sheet('Failed Tests')
    _write_test_rows(failed_ws, failed)
    failed_ws.freeze_panes = 'A2'

    skipped_ws = wb.create_sheet('Skipped Tests')
    _write_test_rows(skipped_ws, skipped)
    skipped_ws.freeze_panes = 'A2'

    _add_metrics_sheet(wb, results)
    _add_defect_summary(wb, failed)

    main_path = os.path.join(output_dir, f'Automation_Test_Report_{suite or "all"}.xlsx')
    wb.save(main_path)
    print(f"Main report: {main_path}")

    # ── 2. Failed tests report ──────────────────────────────────────────────
    if failed:
        wb2 = openpyxl.Workbook()
        wb2.active.title = 'Failed Tests'
        _write_test_rows(wb2.active, failed)
        _add_defect_summary(wb2, failed)
        fail_path = os.path.join(output_dir, 'Failed_Test_Cases.xlsx')
        wb2.save(fail_path)
        print(f"Failed report: {fail_path}")

    # ── 3. Passed tests report ──────────────────────────────────────────────
    if passed:
        wb3 = openpyxl.Workbook()
        wb3.active.title = 'Passed Tests'
        _write_test_rows(wb3.active, passed)
        pass_path = os.path.join(output_dir, 'Passed_Test_Cases.xlsx')
        wb3.save(pass_path)
        print(f"Passed report: {pass_path}")

    # ── 4. Summary report ───────────────────────────────────────────────────
    wb4 = openpyxl.Workbook()
    ws4 = wb4.active
    ws4.title = 'Summary'
    ws4.sheet_view.showGridLines = False
    ws4.column_dimensions['A'].width = 30
    ws4.column_dimensions['B'].width = 20

    total = len(results)
    pass_rate = round(len(passed) / total * 100, 2) if total else 0

    ws4.cell(1, 1, 'FlowTime E2E Test Summary').font = Font(bold=True, size=14, color='1E3A5F')
    summary_data = [
        ('Total Tests', total), ('Passed', len(passed)),
        ('Failed', len(failed)), ('Skipped', len(skipped)),
        ('Pass Rate', f'{pass_rate}%'),
        ('Suite', suite or 'All'),
        ('Date', datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')),
        ('Result', 'PASS ✅' if pass_rate >= 95 else 'FAIL ❌'),
    ]
    for row_i, (k, v) in enumerate(summary_data, 3):
        ws4.cell(row_i, 1, k).font = Font(bold=True)
        cell = ws4.cell(row_i, 2, str(v))
        if k == 'Result':
            cell.font = Font(bold=True, color='166534' if '✅' in str(v) else '991B1B')

    sum_path = os.path.join(output_dir, 'Summary_Report.xlsx')
    wb4.save(sum_path)
    print(f"Summary: {sum_path}")


def main():
    parser = argparse.ArgumentParser(description='Excel Report Generator')
    parser.add_argument('--suite', default='')
    parser.add_argument('--output', default='automation/reports/Excel')
    parser.add_argument('--input', default='automation/reports')
    parser.add_argument('--merge', action='store_true')
    args = parser.parse_args()

    input_dir = args.input if args.merge else args.output.replace('/Excel', '')
    results = _load_results(input_dir)
    if not results:
        print(f"No results found in {input_dir}")
        results = []

    os.makedirs(args.output, exist_ok=True)
    generate_excel_reports(results, args.output, args.suite)


if __name__ == '__main__':
    main()
